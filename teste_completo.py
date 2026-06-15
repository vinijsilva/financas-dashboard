import os
import openpyxl
from extrator import _ler_csv_xp, _ler_csv_nubank, _ler_pdf_bb
from classificador import classificar
from escritor import escrever_na_planilha, escrever_receitas
from collections import Counter, defaultdict

_STOP_WORDS = {'DE', 'DO', 'DA', 'DAS', 'DOS', 'E', 'EM', 'NO', 'NA', 'PARA', 'POR', 'COM', 'OS', 'AS'}

def _palavras_chave(texto):
    return {p for p in str(texto).upper().split() if len(p) >= 4 and p not in _STOP_WORDS}

def _ler_custos_fixos_aba(arquivo_excel, aba_nome):
    """Lê os custos fixos (colunas H-M, linha 3+) da aba do mês e retorna lista de (descricao_upper, valor)."""
    try:
        wb = openpyxl.load_workbook(arquivo_excel, read_only=True, keep_vba=True, data_only=True)
        fixos = []
        if aba_nome in wb.sheetnames:
            ws = wb[aba_nome]
            for row in ws.iter_rows(min_row=3, max_col=13, values_only=True):
                if row[7] is not None and row[8] and row[9]:
                    try:
                        desc = str(row[8]).upper().strip()
                        valor = float(row[9])
                        if desc and valor > 0:
                            fixos.append((desc, valor))
                    except Exception:
                        pass
        wb.close()
        return fixos
    except Exception:
        return []

def _verificar_duplicata_fixo(descricao, valor, custos_fixos):
    """Retorna (True, desc_fixo, val_fixo) se a transação bate com algum custo fixo por palavra-chave + valor."""
    palavras = _palavras_chave(descricao)
    v = abs(float(valor))
    for desc_fixo, val_fixo in custos_fixos:
        if abs(v - val_fixo) > 0.02:
            continue
        if palavras & _palavras_chave(desc_fixo):
            return True, desc_fixo, val_fixo
    return False, None, None

PASTA_XP     = r"G:\Meu Drive\Finanças\Faturas\XP"
PASTA_BB     = r"G:\Meu Drive\Finanças\Extratos\2026\BB"
PASTA_NUBANK = r"G:\Meu Drive\Finanças\Extratos\2026\Nubank"
ARQUIVO_EXCEL = r"G:\Meu Drive\Finanças\Finanças 2026.xlsm"

def processar_tudo():
    arquivos_para_processar = []

    for raiz, _, arquivos in os.walk(PASTA_XP):
        for a in arquivos:
            if a.endswith(".csv") and ".processado" not in a:
                arquivos_para_processar.append((os.path.join(raiz, a), "XP"))

    for raiz, _, arquivos in os.walk(PASTA_BB):
        for a in arquivos:
            if a.endswith(".pdf") and ".processado" not in a:
                arquivos_para_processar.append((os.path.join(raiz, a), "BB"))

    for raiz, _, arquivos in os.walk(PASTA_NUBANK):
        for a in arquivos:
            if a.endswith(".csv") and ".processado" not in a:
                arquivos_para_processar.append((os.path.join(raiz, a), "NUBANK"))

    if not arquivos_para_processar:
        print("\n[INFO] Nenhum arquivo novo encontrado para processar.")
        return

    # Coleta todas as transações de todos os arquivos antes de escrever
    despesas_por_aba = defaultdict(list)  # {aba: [transacoes]}
    todas_receitas = []
    arquivos_processados = []

    mapa = {1:'Jan', 2:'Fev', 3:'Mar', 4:'Abr', 5:'Mai', 6:'Jun',
            7:'Jul', 8:'Ago', 9:'Set', 10:'Out', 11:'Nov', 12:'Dez'}

    for caminho_arquivo, origem in arquivos_para_processar:
        nome_arquivo = os.path.basename(caminho_arquivo)
        print(f"\n--- Lendo [{origem}]: {nome_arquivo} ---")

        try:
            if origem == "XP":      transacoes = _ler_csv_xp(caminho_arquivo)
            elif origem == "NUBANK": transacoes = _ler_csv_nubank(caminho_arquivo)
            elif origem == "BB":    transacoes = _ler_pdf_bb(caminho_arquivo)

            if not transacoes:
                print(f"-> Nenhuma transação válida extraída.")
                continue

            meses_filtrados = [t['data'].month for t in transacoes if t['parcela'] == '-' and "Pagamento" not in t['descricao']]
            anos_filtrados  = [t['data'].year  for t in transacoes if t['parcela'] == '-' and "Pagamento" not in t['descricao']]

            if not meses_filtrados:
                meses_filtrados = [t['data'].month for t in transacoes]
                anos_filtrados  = [t['data'].year  for t in transacoes]

            mes_pred = Counter(meses_filtrados).most_common(1)[0][0]
            ano_pred = Counter(anos_filtrados).most_common(1)[0][0]
            aba_destino = f"{mapa[mes_pred]}.{ano_pred}"

            for t in transacoes:
                try:
                    resultado = classificar(t['descricao'], t['portador'])
                    cat = resultado[0] if isinstance(resultado, tuple) else resultado
                except Exception:
                    cat = "Outros"

                t['origem'] = origem
                t['categoria'] = cat

                if t['valor'] < 0:
                    t['valor'] = abs(t['valor'])
                    despesas_por_aba[aba_destino].append(t)
                elif t['valor'] > 0 and origem != 'XP':
                    # XP é fatura de cartão: créditos/estornos não são receita real
                    t['valor'] = abs(t['valor'])
                    todas_receitas.append(t)

            arquivos_processados.append(caminho_arquivo)

        except Exception as e:
            print(f"Erro crítico ao processar {nome_arquivo}: {e}")

    # Verifica duplicatas com os custos fixos da aba de destino antes de escrever
    for aba in list(despesas_por_aba.keys()):
        custos_fixos = _ler_custos_fixos_aba(ARQUIVO_EXCEL, aba)
        if not custos_fixos:
            continue
        filtradas = []
        for t in despesas_por_aba[aba]:
            duplicata, desc_fixo, val_fixo = _verificar_duplicata_fixo(t['descricao'], t['valor'], custos_fixos)
            if duplicata:
                print(f"  [DUPLICATA] '{t['descricao']}' R${t['valor']:.2f} já está nos custos fixos como '{desc_fixo}' R${val_fixo:.2f} — ignorado.")
            else:
                filtradas.append(t)
        pulados = len(despesas_por_aba[aba]) - len(filtradas)
        if pulados:
            print(f"  >> {pulados} lancamento(s) ignorado(s) por ja constarem nos custos fixos de '{aba}'.")
        despesas_por_aba[aba] = filtradas

    # Escreve despesas por aba, todas ordenadas por data
    for aba, despesas in despesas_por_aba.items():
        despesas.sort(key=lambda t: t['data'])
        escrever_na_planilha(ARQUIVO_EXCEL, aba, despesas)
        print(f"\nSucesso: {len(despesas)} saídas lançadas na aba '{aba}'")

    if todas_receitas:
        escrever_receitas(ARQUIVO_EXCEL, todas_receitas)
        print(f"Sucesso: {len(todas_receitas)} entradas lançadas na aba 'Recebimentos'")

    # Renomeia arquivos apenas após escrita bem-sucedida
    for caminho in arquivos_processados:
        os.rename(caminho, caminho + ".processado")

if __name__ == "__main__":
    processar_tudo()
