# -*- coding: utf-8 -*-
import os
import openpyxl
from collections import defaultdict
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')
import google.generativeai as genai
import ollama as _ollama

PLANILHA = r"G:\Meu Drive\Finanças\Finanças 2026.xlsm"

# Carrega .env se existir
_env = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(_env):
    for _linha in open(_env):
        if '=' in _linha and not _linha.startswith('#'):
            _k, _v = _linha.strip().split('=', 1)
            os.environ.setdefault(_k.strip(), _v.strip())

# Cobranças mensais do cartão XP que já estão nos custos fixos.
# Adicione aqui outros itens recorrentes que aparecem duplicados.
_FIXOS_CARTAO_XP = [
    ('MARATHON', 189.00),  # academia
    ('APPLE',      5.90),  # Apple.bill
]

def _eh_duplicata_fixo(descricao, valor, cartao):
    """True se a transação variável já está contabilizada nos custos fixos."""
    if 'XP' not in str(cartao).upper():
        return False
    d = str(descricao).upper()
    v = float(valor)
    return any(nome in d and abs(v - val_ref) < 0.02 for nome, val_ref in _FIXOS_CARTAO_XP)


MESES_ABAS = [
    'Jan.2026', 'Fev.2026', 'Mar.2026', 'Abr.2026',
    'Mai.2026', 'Jun.2026', 'Jul.2026', 'Ago.2026',
    'Set.2026', 'Out.2026', 'Nov.2026', 'Dez.2026',
]

_MES_NUM = {
    'Jan': 1, 'Fev': 2, 'Mar': 3, 'Abr': 4,
    'Mai': 5, 'Jun': 6, 'Jul': 7, 'Ago': 8,
    'Set': 9, 'Out': 10, 'Nov': 11, 'Dez': 12,
}

def _fmt_brl(v):
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def ler_dados():
    wb = openpyxl.load_workbook(PLANILHA, read_only=True, keep_vba=True, data_only=True)

    # --- Despesas por mês ---
    despesas = {}   # {aba: [{'data', 'descricao', 'valor', 'cartao', 'categoria', 'parcela', 'tipo'}]}
    total_mes = {}  # {aba: float}

    for aba in MESES_ABAS:
        if aba not in wb.sheetnames:
            continue
        ws = wb[aba]
        linhas = []

        partes = aba.split('.')
        mes_num = _MES_NUM.get(partes[0], 1)
        ano_num = int(partes[1]) if len(partes) > 1 else 2026

        for row in ws.iter_rows(min_row=3, max_col=13, values_only=True):
            # Custos variáveis (A-F, índices 0-5)
            if row[0]:
                try:
                    data = row[0].strftime('%d/%m/%Y') if hasattr(row[0], 'strftime') else str(row[0])[:10]
                    valor = float(row[2]) if row[2] else 0.0
                    descricao = str(row[1] or '')
                    cartao_val = str(row[3] or '')
                    if _eh_duplicata_fixo(descricao, valor, cartao_val):
                        continue
                    linhas.append({
                        'data': data,
                        'descricao': descricao,
                        'valor': valor,
                        'cartao': cartao_val,
                        'categoria': str(row[4] or ''),
                        'parcela': str(row[5] or ''),
                        'tipo': 'variavel',
                    })
                except Exception:
                    pass

            # Custos fixos (H-M, índices 7-12): dia, descrição, valor, cartão, categoria, parcela
            if row[7]:
                try:
                    dia = int(float(row[7]))
                    import calendar
                    dia = min(dia, calendar.monthrange(ano_num, mes_num)[1])
                    data = datetime(ano_num, mes_num, dia).strftime('%d/%m/%Y')
                    valor = float(row[9]) if row[9] else 0.0
                    if not valor:
                        continue
                    linhas.append({
                        'data': data,
                        'descricao': str(row[8] or ''),
                        'valor': valor,
                        'cartao': str(row[10] or ''),
                        'categoria': str(row[11] or ''),
                        'parcela': str(row[12] or ''),
                        'tipo': 'fixo',
                    })
                except Exception:
                    pass

        if linhas:
            despesas[aba] = linhas
            total_mes[aba] = sum(l['valor'] for l in linhas)

    # --- Receitas ---
    receitas = []
    if 'Recebimentos' in wb.sheetnames:
        ws = wb['Recebimentos']
        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            if not row[0] or not row[6]:
                continue
            try:
                data = row[3].strftime('%d/%m/%Y') if hasattr(row[3], 'strftime') else str(row[3])[:10]
                receitas.append({
                    'origem': str(row[0]),
                    'categoria': str(row[1] or ''),
                    'referencia': str(row[2] or ''),
                    'data': data,
                    'tipo': str(row[5] or ''),
                    'valor': float(row[6]),
                })
            except Exception:
                continue

    wb.close()
    return despesas, total_mes, receitas


def montar_contexto(despesas, total_mes, receitas):
    linhas = []
    linhas.append("=== PLANILHA FINANCEIRA 2026 — DADOS COMPLETOS ===\n")

    # Resumo mensal de despesas
    linhas.append("## DESPESAS POR MÊS\n")
    for aba, transacoes in despesas.items():
        total = total_mes[aba]
        linhas.append(f"### {aba} — Total: {_fmt_brl(total)}")

        # Agrupa por categoria
        por_cat = defaultdict(float)
        for t in transacoes:
            por_cat[t['categoria']] += t['valor']
        for cat, val in sorted(por_cat.items(), key=lambda x: -x[1]):
            linhas.append(f"  {cat}: {_fmt_brl(val)}")
        linhas.append("")

    # Detalhamento de transações por mês
    linhas.append("\n## TRANSAÇÕES DETALHADAS POR MÊS\n")
    for aba, transacoes in despesas.items():
        linhas.append(f"### {aba}")
        for t in transacoes:
            parcela = f" [{t['parcela']}]" if t['parcela'] and t['parcela'] != '-' else ''
            linhas.append(
                f"  {t['data']} | {t['descricao'][:40]:<40} | {_fmt_brl(t['valor']):>12} | "
                f"{t['cartao']:<14} | {t['categoria']}{parcela}"
            )
        linhas.append("")

    # Receitas
    linhas.append("\n## RECEBIMENTOS\n")
    por_tipo_rec = defaultdict(float)
    for r in receitas:
        por_tipo_rec[r['categoria']] += r['valor']
    for cat, val in sorted(por_tipo_rec.items(), key=lambda x: -x[1]):
        linhas.append(f"  {cat}: {_fmt_brl(val)}")
    linhas.append("")

    # Receitas detalhadas dos últimos meses
    rec_2026 = [r for r in receitas if '2026' in r['data']]
    if rec_2026:
        linhas.append("### Recebimentos em 2026 (detalhado)")
        for r in sorted(rec_2026, key=lambda x: x['data']):
            linhas.append(
                f"  {r['data']} | {r['origem']:<20} | {r['categoria']:<25} | "
                f"{_fmt_brl(r['valor']):>12} | {r['tipo']}"
            )

    return "\n".join(linhas)


SYSTEM_PROMPT = """Você é um assistente financeiro pessoal inteligente e direto.
Você tem acesso aos dados completos da planilha financeira do usuário para 2026.

Categorias de despesas usadas:
- MORADIA: Supermercado, Manutenção/Reforma, Outros
- CARRO: Seguro, Combustível, IPVA/Licenciamento, Lavagem, Manutenção, Outros
- MOTO: Seguro, Combustível, IPVA/Licenciamento, Lavagem, Manutenção, Outros
- BARCO: Gasolina, Óleo, Guardaria, Manutenção, Outros
- D. PESSOAL: Esporte, Educação, Entretenimento, Restaurante, Saúde, Vestuário, Viagens, Outros
- INVEST.: Renda Fixa, Ações, Fundos Imobiliários, Títulos Públicos, Casal
- THALI: PIX, Calvin (pet), Casa, Outros (gastos relacionados à Thalissa)
- JULI: Heitor, Psicóloga, Casa, Outros (gastos relacionados à Júlia)
- V2PA / SAFO ENG.: empresas do usuário (contador, impostos, PIX)

Coluna D (origem do gasto): XP (cartão de crédito XP), PIX - NUBANK, PIX - BB, NUBANK.

Ao responder:
- Use valores em Reais (R$) formatados com vírgula decimal
- Compare meses quando relevante
- Dê insights práticos e diretos
- Se o usuário pedir previsões ou tendências, baseie-se nos dados disponíveis
- Responda sempre em português brasileiro

Dados da planilha:
{dados}
"""


def main():
    api_key = os.environ.get('GEMINI_API_KEY', '').strip()
    if not api_key:
        print("=" * 60)
        print("Para usar o agente, você precisa de uma API key gratuita.")
        print("Acesse: https://aistudio.google.com/apikey")
        print("=" * 60)
        api_key = input("Cole sua GEMINI_API_KEY aqui: ").strip()
        if not api_key:
            print("API key não fornecida. Encerrando.")
            return

    print("\nCarregando dados da planilha...")
    try:
        despesas, total_mes, receitas = ler_dados()
    except Exception as e:
        print(f"Erro ao ler planilha: {e}")
        return

    meses_com_dados = list(despesas.keys())
    total_geral = sum(total_mes.values())

    print(f"Dados carregados: {len(meses_com_dados)} meses, {sum(len(v) for v in despesas.values())} transações")
    print(f"Meses disponíveis: {', '.join(meses_com_dados)}")
    print(f"Total gasto em 2026: {_fmt_brl(total_geral)}")
    print()

    contexto = montar_contexto(despesas, total_mes, receitas)

    system_prompt_completo = SYSTEM_PROMPT.format(dados=contexto)
    usar_gemini = bool(api_key)
    gemini_chat = None

    if usar_gemini:
        try:
            genai.configure(api_key=api_key)
            gmodel = genai.GenerativeModel(
                model_name='gemini-2.5-flash',
                system_instruction=system_prompt_completo,
            )
            # Teste rápido
            gmodel.generate_content('ok')
            gemini_chat = gmodel.start_chat(history=[])
            print("Modelo: Gemini 2.5 Flash")
        except Exception as e:
            print(f"Gemini indisponível ({e.__class__.__name__}), usando ollama/llama3.")
            usar_gemini = False

    if not usar_gemini:
        print("Modelo: ollama / llama3 (local)")

    historico_ollama = [{'role': 'system', 'content': system_prompt_completo}]

    print("=" * 60)
    print("Agente financeiro pronto! Digite sua pergunta.")
    print("(Digite 'sair' para encerrar)")
    print("=" * 60)

    while True:
        try:
            pergunta = input("\nVocê: ").strip()
        except (KeyboardInterrupt, EOFError):
            break

        if not pergunta:
            continue
        if pergunta.lower() in ('sair', 'exit', 'quit', 'q'):
            print("Encerrando agente.")
            break

        if usar_gemini and gemini_chat:
            try:
                resposta = gemini_chat.send_message(pergunta)
                print(f"\nAgente: {resposta.text}")
                continue
            except Exception as e:
                print(f"(Gemini falhou: {e.__class__.__name__}, usando ollama)")
                usar_gemini = False

        # Fallback: ollama
        historico_ollama.append({'role': 'user', 'content': pergunta})
        try:
            r = _ollama.chat(model='llama3', messages=historico_ollama)
            texto = r['message']['content']
            historico_ollama.append({'role': 'assistant', 'content': texto})
            print(f"\nAgente: {texto}")
        except Exception as e:
            historico_ollama.pop()
            print(f"\nErro ao consultar ollama: {e}")


if __name__ == "__main__":
    main()
